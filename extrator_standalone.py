import os
import re
import sys
import time
import json
import threading
import subprocess
import requests
import fitz  # PyMuPDF
import pandas as pd
import customtkinter as ctk
import ctypes  # << Necessário para forçar o ícone na barra de tarefas do Windows
import concurrent.futures  # Multi-threading de alta performance
import tkinter as tk
import queue  # Fila thread-safe para logs
import unicodedata  # Normalização Unicode para deduplicação
from datetime import datetime  # Timestamps nos logs
from tkinter import filedialog, messagebox

# ==========================================
# CONFIGURAÇÕES DO SUPABASE / SEGURANÇA
# ==========================================
SUPABASE_URL = "https://qhjyvdywvdcbdgicwjzu.supabase.co/rest/v1"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFoanl2ZHl3dmRjYmRnaWN3anp1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzE1MzU4NDUsImV4cCI6MjA4NzExMTg0NX0.d9wzf3laY6suarJh_InWij_JjOUcqCW-5mpDAKJuC68"

# ==========================================
# ARQUIVO LOCAL DE CONFIGURAÇÃO (SALVAR CHAVES)
# ==========================================
CONFIG_FILE = "config_extrator.json"

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return {"gemini_api_key": "", "license_key": ""}

def save_config(config_data):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config_data, f)

# ==========================================
# FUNÇÕES DE HARDWARE ID E SEGURANÇA
# ==========================================
def get_hardware_id():
    """Gera um identificador único baseada na placa-mãe/processador (Apenas Windows)"""
    try:
        # Tenta PowerShell primeiro (recomendado para Windows 11 que depreciou o wmic nativo)
        cmd = ['powershell', '-NoProfile', '-Command', '(Get-CimInstance -Class Win32_ComputerSystemProduct).UUID']
        output = subprocess.check_output(cmd, stderr=subprocess.DEVNULL).decode().strip()
        if output:
            return output
    except Exception:
        pass
        
    try:
        # Fallback de segurança CMD antigo, mas sem poluir o console com mensagem vermelha
        output = subprocess.check_output('wmic csproduct get uuid', shell=True, stderr=subprocess.DEVNULL).decode().strip().split('\n')
        hwid = output[1].strip()
        if hwid:
            return hwid
    except Exception:
        pass
        
    # Ultimo recurso se todas as permissões do Windows falharem (Usa a Placa de Rede MAC)
    import uuid
    return str(uuid.getnode())

def validate_license(license_key):
    """Bate no Supabase, verifica se a licença é válida e casa com o Computador Atual."""
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    
    # 1. Busca a Licença
    resp = requests.get(f"{SUPABASE_URL}/licenses?license_key=eq.{license_key}", headers=headers)
    if resp.status_code != 200:
        return False, f"Erro de conexão com servidor de licenças. {resp.status_code}"
    
    data = resp.json()
    if not data:
        return False, "Licença não encontrada."
        
    lic = data[0]
    
    if not lic.get("is_active"):
        return False, "Sua licença foi desativada/bloqueada. Contate o suporte."
        
    current_hwid = get_hardware_id()
    db_hwid = lic.get("hardware_id")
    
    # Se a licença é "Virgem" (nunca ativada), casamos ela com a máquina de agora!
    if db_hwid is None or db_hwid.strip() == "":
        patch_resp = requests.patch(
            f"{SUPABASE_URL}/licenses?license_key=eq.{license_key}", 
            headers=headers, 
            json={"hardware_id": current_hwid}
        )
        if patch_resp.status_code == 200:
            return True, f"Licença ativada c/ Sucesso nesta máquina! Olá, {lic.get('client_name')}!"
        else:
            return False, "Falha ao vincular sua licença na primeira ativação."
            
    # Se a Licença já tem um PC cadastrado, checa se é igual ao dono...
    if db_hwid != current_hwid:
        return False, "Essa licença já foi ativada em outro computador.\nO acesso não é permitido em múltiplas máquinas."
        
    return True, f"Login Autorizado. Bem-vindo, {lic.get('client_name')}"


# ==========================================
# CÓDIGO DO MOTOR DE EXTRAÇÃO E INTELIGÊNCIA
# ==========================================
SYSTEM_PROMPT = """
Você é um robô puramente extrativista e formatador de anais de congressos e periódicos acadêmicos em PDF. Sua única missão: extrair TÍTULO DO TRABALHO, NOME DOS AUTORES e E-MAIL de cada autor, e retornar EXCLUSIVAMENTE uma tabela Markdown.

=== PADRÃO DO DOCUMENTO (MUITO IMPORTANTE) ===
Os documentos seguem layouts variados. Procure em TODOS estes locais:
  1. TÍTULO DO TRABALHO: em MAIÚSCULAS e/ou negrito, centralizado no topo da 1ª página do artigo.
  2. AUTORES: logo abaixo do título, com números sobrescritos (ex: "Maria Silva¹, João Santos²").
  3. E-MAILS podem estar em QUALQUER um destes locais:
     a) RODAPÉ da mesma página (com sobrescritos: "¹ Mestranda, UEFS. maria@uefs.br")
     b) Bloco de AFILIAÇÃO logo após os nomes dos autores
     c) Seção "Corresponding author" ou "Autor correspondente" ou "E-mail:" ou "Contact:"
     d) Rodapé com asterisco (*) indicando autor de correspondência
     e) Linha com formato "Nome Sobrenome (email@dominio.br)"
     f) Blocos biográficos no final do artigo ("Sobre os autores")
     g) Cabeçalhos ou rodapés repetidos com informação de contato

=== REGRA DE OURO — UMA LINHA POR AUTOR ===
Se um trabalho tiver 4 autores, você DEVE gerar 4 linhas na tabela — uma para cada autor.
O título do trabalho deve ser REPETIDO em cada linha. NUNCA agrupe autores na mesma célula.
NUNCA OMITA UM AUTOR. Se há 5 nomes listados, devem haver 5 linhas.

=== CAÇA AO E-MAIL (OBSESSIVA) ===
1. Percorra CADA LINHA do texto procurando por padrões contendo "@".
2. Associe e-mails aos autores pelo índice numérico sobrescrito (¹²³⁴ ou 1.2.3.4 ou (1)(2) ou *).
3. Se um e-mail aparece isolado sem índice, associe-o ao autor mais próximo ou ao primeiro autor.
4. Se houver apenas 1 e-mail para múltiplos autores, coloque-o no primeiro autor e deixe os demais vazios.
5. Se o e-mail NÃO aparece neste trecho, deixe a célula VAZIA (não invente).

=== REGRAS OBRIGATÓRIAS DE FORMATAÇÃO ===
1. Cabeçalho EXATO: `| Títulos dos Trabalhos | Nomes dos Autores | E-mails dos Autores |`
2. Linha separadora EXATA: `|---|---|---|`
3. REMOVA quebras de linha (Enter) de dentro do TÍTULO. O título deve ser uma linha contínua.
4. LIMPEZA de nomes: remova números de afiliação (ex: "Maria Silva1" → "Maria Silva").
5. Ignore sumários, capas, referências bibliográficas e agradecimentos.
6. Sua resposta deve conter SOMENTE a tabela Markdown — sem texto antes ou depois.
"""

# Cache global do modelo Gemini (P1: elimina requests de descoberta repetidos)
_gemini_model_cache = {}
_model_blacklist: set = set()  # modelos que retornaram 404 em runtime

# Rate limiting proativo: máx 3 requests simultâneos + 4.2s de espaçamento.
# Garante 14.2 Requests Por Minuto (RPM), o que é abaixo do limite Free (15 RPM)
# Assim, o cliente pode até usar de graça sem gastar nem os 23 reais.
_api_semaphore = threading.Semaphore(3)
_API_SPACING = 4.2  

def _get_gemini_model(api_key):
    """Descobre o melhor modelo Flash UMA VEZ e cacheia."""
    if api_key in _gemini_model_cache and _gemini_model_cache[api_key] in _model_blacklist:
        del _gemini_model_cache[api_key]

    if api_key in _gemini_model_cache:
        return _gemini_model_cache[api_key]
    
    models_url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
    resp_models = requests.get(models_url, timeout=15)
    if resp_models.status_code != 200:
        raise Exception(f"Erro ao validar a chave API (HTTP {resp_models.status_code})")
        
    available_models = [
        m.get('name') for m in resp_models.json().get('models', [])
        if 'generateContent' in m.get('supportedGenerationMethods', [])
        and m.get('name') not in _model_blacklist
    ]
    
    # Prioridade focada TOTAL na economia: Flash Lite é de 50% a 70% mais barato!
    PREFERRED = [
        "models/gemini-2.0-flash-lite-preview-02-05", # O modelo mais barato e rápido do Google
        "models/gemini-2.0-flash-lite",               # Fallback para versão final do Lite
        "models/gemini-1.5-flash-8b",                 # Modelo 1.5 super otimizado para economia extrema
        "models/gemini-1.5-flash",                    # O 1.5 Flash clássico, excelente e hiper barato
        "models/gemini-2.0-flash",                    # Flash 2.0
        "models/gemini-flash-latest"                  # Genérico
    ]
    target_model = None
    for pref in PREFERRED:
        if pref in available_models:
            target_model = pref
            break
            
    if not target_model:
        flash_models = [m for m in available_models if 'flash' in m.lower()]
        if flash_models:
            target_model = flash_models[0]
        else:
            raise Exception(f"Nenhum modelo Flash encontrado na chave: {available_models}")
    
    _gemini_model_cache[api_key] = target_model
    return target_model

def extract_from_gemini(text_content, api_key):
    target_model = _get_gemini_model(api_key)  # P1: Usa cache
    url = f"https://generativelanguage.googleapis.com/v1beta/{target_model}:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    
    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": [{"parts": [{"text": f"Extraia os trabalhos deste texto:\n\n{text_content}"}]}],
        "generationConfig": {
            "temperature": 0.0,       # determinístico — extração não precisa de criatividade
            "maxOutputTokens": 65536, # margem ampla para tabelas com muitos artigos/autores
        }
    }
    
    # Timeout de 400 segundos: blocos de Índice/Sumário contam com CENTENAS de trabalhos
    # e a IA pode levar 3 ou 4 minutos inteiros só para formatar essa saída gigante no modelo 1.5.
    response = requests.post(url, headers=headers, json=payload, timeout=400)
    if response.status_code == 404:
        # Modelo indisponível para esta chave — adiciona à blacklist e força redescoberta
        _model_blacklist.add(target_model)
        _gemini_model_cache.pop(api_key, None)
        raise Exception(
            f"Modelo {target_model} retornou 404 (indisponível nesta chave). "
            f"Tentando próximo modelo na redescoberta..."
        )
    if response.status_code != 200:
        raise Exception(f"Erro da API do Gemini ({target_model}) (HTTP {response.status_code}): {response.text}")
        
    data = response.json()
    try:
        text = data['candidates'][0]['content']['parts'][0]['text']
        # Detecta truncamento: se o modelo parou por limite de tokens, a tabela está incompleta
        finish_reason = data['candidates'][0].get('finishReason', '')
        if finish_reason == 'MAX_TOKENS':
            text += "\n<!-- TRUNCADO: resposta cortada pelo limite de tokens -->"
        return text
    except (KeyError, IndexError):
        raise Exception(f"Resposta inesperada do Gemini:\n{json.dumps(data, indent=2)}")

def extract_from_openai(text_content, api_key):
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }
    
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {
                "role": "system",
                "content": SYSTEM_PROMPT
            },
            {
                "role": "user",
                "content": f"Extraia os trabalhos deste texto:\n\n{text_content}"
            }
        ],
        "temperature": 0.0
    }
    
    response = requests.post(url, headers=headers, json=payload, timeout=120)
    if response.status_code != 200:
        raise Exception(f"Erro da API da OpenAI (HTTP {response.status_code}): {response.text}")
        
    data = response.json()
    try:
        return data['choices'][0]['message']['content']
    except (KeyError, IndexError):
        raise Exception(f"Resposta inesperada da OpenAI:\n{json.dumps(data, indent=2)}")

def parse_markdown_table_to_dicts(markdown_text):
    """Converte a string de tabela Markdown retornada pela LLM em JSON.
    R4: Parsing defensivo — detecta cabeçalho vs dados automaticamente."""
    linhas = markdown_text.split('\n')
    linhas = [linha for linha in linhas if linha.strip() and '|' in linha]
    
    # Remove linhas separadoras (---|---|---)
    linhas_dados = [l for l in linhas if not l.replace(' ', '').replace('|', '').replace('-', '') == '']
    if not linhas_dados:
        return []
        
    chaves = ["Títulos dos Trabalhos", "Nomes dos Autores", "E-mails dos Autores"]
    
    # R4: Detecta se a primeira linha é cabeçalho (Verificação rígida para evitar falso positivo)
    primeira_linha_cols = [col.strip().lower() for col in linhas_dados[0].split('|') if col.strip()]
    is_header = False
    if len(primeira_linha_cols) > 0 and ('título' in primeira_linha_cols[0] or 'titulo' in primeira_linha_cols[0] or 'trabalho' in primeira_linha_cols[0]):
        is_header = True
        
    linhas_conteudo = linhas_dados[1:] if is_header else linhas_dados
    
    resultados = []
    for linha in linhas_conteudo:
        cols = [col.strip() for col in linha.split('|')][1:-1]
        
        if len(cols) >= 3:
            email_raw = cols[2]
            # M2: Extração de e-mails inclusiva e segura
            email_limpo = email_raw.strip()
            if email_limpo and '@' in email_limpo:
                emails_encontrados = re.findall(r'[\w.+-]+@[\w.-]+\.\w{2,}', email_limpo)
                email_limpo = ", ".join(emails_encontrados) if emails_encontrados else ""
            
            obj = {
                chaves[0]: cols[0],
                chaves[1]: cols[1],
                chaves[2]: email_limpo
            }
            resultados.append(obj)
            
    return resultados


# ==========================================
# FUNÇÕES DE PÓS-PROCESSAMENTO
# ==========================================
import re as _re

def consolidar_projetos(projetos):
    """
    Merge inteligente: une linhas com mesmo título+autor,
    priorizando o e-mail preenchido quando chunks diferentes
    capturam a mesma pessoa em momentos distintos.
    R3: Normalização Unicode para evitar duplicatas por encoding diferente.
    """
    mapa = {}
    chave_titulo = "Títulos dos Trabalhos"
    chave_autor  = "Nomes dos Autores"
    chave_email  = "E-mails dos Autores"
    
    for proj in projetos:
        titulo = (proj.get(chave_titulo) or "").strip()
        autor  = (proj.get(chave_autor)  or "").strip()
        email  = (proj.get(chave_email)  or "").strip()
        # R3: Normaliza Unicode (NFC) para deduplicação correta de acentos
        titulo_norm = unicodedata.normalize('NFC', titulo.lower())
        autor_norm = unicodedata.normalize('NFC', autor.lower())
        chave_composta = f"{titulo_norm}|||{autor_norm}"
        
        if chave_composta not in mapa:
            mapa[chave_composta] = {chave_titulo: titulo, chave_autor: autor, chave_email: email}
        else:
            if email and not mapa[chave_composta][chave_email]:
                mapa[chave_composta][chave_email] = email
    
    return list(mapa.values())


def expandir_por_autor(projetos):
    """
    Caso a IA retorne múltiplos autores agrupados numa célula
    (ex: 'Maria Silva, João Santos'), desdobra em linhas individuais.
    Remove números/símbolos sobrescritos dos nomes (¹²³⁴⁵⁶⁷⁸⁹⁰ e variantes).
    """
    chave_titulo = "Títulos dos Trabalhos"
    chave_autor  = "Nomes dos Autores"
    chave_email  = "E-mails dos Autores"
    sobrescritos = str.maketrans("", "", "¹²³⁴⁵⁶⁷⁸⁹⁰₁₂₃₄₅₆₇₈₉₀")
    
    expandido = []
    for proj in projetos:
        titulo  = (proj.get(chave_titulo) or "").strip()
        autores = (proj.get(chave_autor)  or "").strip()
        emails  = (proj.get(chave_email)  or "").strip()
        
        # Remove sobrescritos dos nomes
        autores_limpos = autores.translate(sobrescritos).strip()
        
        # Separa múltiplos autores (separados por vírgula ou ponto e vírgula)
        lista_autores = [a.strip() for a in _re.split(r"[;,]", autores_limpos) if a.strip()]
        # Separa múltiplos e-mails na mesma célula
        lista_emails  = [e.strip() for e in _re.split(r"[;,\s]+", emails) if "@" in e]
        
        if len(lista_autores) <= 1:
            # Já está no formato certo (1 autor por linha)
            expandido.append({
                chave_titulo: titulo,
                chave_autor:  autores_limpos or autores,
                chave_email:  emails
            })
        else:
            # Desdobra: um registro por autor
            for i, autor in enumerate(lista_autores):
                email_autor = lista_emails[i] if i < len(lista_emails) else ""
                expandido.append({
                    chave_titulo: titulo,
                    chave_autor:  autor,
                    chave_email:  email_autor
                })
    
    return expandido

# ==========================================
# FUNÇÃO DE SALVAMENTO COM FORMATAÇÃO
# ==========================================
def salvar_excel_formatado(projetos, filepath):
    """
    Salva a lista de projetos em Excel com:
    - Largura de coluna automática (ajusta ao conteúdo mais longo)
    - Cabeçalho em negrito
    - Fallback automático para arquivo de backup se o original estiver aberto
    Lança PermissionError com mensagem amigável se ambos falharem.
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    colunas = ["Títulos dos Trabalhos", "Nomes dos Autores", "E-mails dos Autores"]

    # Cabeçalho em negrito
    for col_idx, nome_col in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.font = Font(bold=True)

    # Dados
    for row_idx, proj in enumerate(projetos, start=2):
        ws.cell(row=row_idx, column=1, value=proj.get(colunas[0], ""))
        ws.cell(row=row_idx, column=2, value=proj.get(colunas[1], ""))
        ws.cell(row=row_idx, column=3, value=proj.get(colunas[2], ""))

    # Ajuste automático de largura de coluna
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    # Tenta salvar no caminho original; fallback para _backup se aberto
    try:
        wb.save(filepath)
    except PermissionError:
        backup_path = filepath.replace(".xlsx", "_backup.xlsx")
        try:
            wb.save(backup_path)
            raise PermissionError(
                f"Arquivo Excel estava aberto. Dados salvos em:\n{backup_path}"
            )
        except Exception as e2:
            raise PermissionError(f"Não foi possível salvar o Excel: {e2}")

# ==========================================
# BOTÃO 3D METÁLICO (Canvas-based)
# ==========================================
class MetalButton3D(tk.Canvas):
    """Botão com efeito 3D metálico usando Canvas do tkinter.
    Simula gradiente, relevo e texto prateado."""
    
    # Paleta de cores para cada estado
    _STATES = {
        "normal":   {"top": "#1de9a0", "mid": "#0ecf85", "bot": "#0a9460", "shadow": "#065c3c", "text": "#d0ffe8", "border": "#0a7a50"},
        "hover":    {"top": "#26ffb3", "mid": "#18e895", "bot": "#0db870", "shadow": "#086b42", "text": "#ffffff", "border": "#0db870"},
        "pressed":  {"top": "#0a9460", "mid": "#087a50", "bot": "#065c3c", "shadow": "#033520", "text": "#a0ffd4", "border": "#044d30"},
        "disabled": {"top": "#3a6655", "mid": "#2e5244", "bot": "#1e3a2f", "shadow": "#111f18", "text": "#6aad8e", "border": "#1e3a2f"},
    }
    
    def __init__(self, master, text="", command=None, height=52, **kwargs):
        super().__init__(master, height=height, bd=0, highlightthickness=0,
                         cursor="hand2", bg="#1a1a2e", **kwargs)
        self._text    = text
        self._command = command
        self._state   = "normal"
        self._height  = height
        self._shadow_h = 5  # espessura da sombra inferior (efeito Z)
        
        self.bind("<Configure>",    self._redraw)
        self.bind("<Enter>",        lambda e: self._set_hover(True))
        self.bind("<Leave>",        lambda e: self._set_hover(False))
        self.bind("<ButtonPress-1>",  lambda e: self._set_pressed(True))
        self.bind("<ButtonRelease-1>",self._on_release)
    
    def _palette(self):
        return self._STATES[self._state]
    
    def _redraw(self, event=None):
        self.delete("all")
        w = self.winfo_width()
        h = self._height
        r = 10  # raio do border-radius
        sh = self._shadow_h
        p = self._palette()
        
        # --- Camada de sombra (base inferior) ---
        self._rounded_rect(0, sh, w, h, r, fill=p["shadow"], outline="")
        
        # --- Corpo principal (simulando gradiente em 3 faixas horizontais) ---
        # Faixa superior (\u00e9 o reflexo de luz)
        self._rounded_rect(0, 0, w, h - sh, r, fill=p["top"], outline="")
        # Faixa m\u00e9dia
        band_y = int((h - sh) * 0.38)
        self._rounded_rect(0, band_y, w, h - sh, r, fill=p["mid"], outline="")
        # Faixa inferior escura (cria a profundidade)
        self._rounded_rect(0, int((h - sh) * 0.68), w, h - sh, r, fill=p["bot"], outline="")
        
        # --- Borda interna (realce 1px no topo) ---
        self._rounded_rect(1, 1, w - 1, h - sh - 1, r - 1, fill="",
                           outline="#3dffc0" if self._state not in ("disabled", "pressed") else p["border"],
                           width=1)
        
        # --- Linha de borda externa ---
        self._rounded_rect(0, 0, w, h - sh, r, fill="", outline=p["border"], width=1)
        
        # --- Texto met\u00e1lico com sombra ---
        cx, cy = w // 2, (h - sh) // 2
        # Sombra do texto
        self.create_text(cx + 1, cy + 2, text=self._text,
                         font=("Segoe UI", 13, "bold"),
                         fill="#003322", anchor="center")
        # Texto principal (cor prateada/metálica)
        self.create_text(cx, cy, text=self._text,
                         font=("Segoe UI", 13, "bold"),
                         fill=p["text"], anchor="center")
    
    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        """Desenha um retângulo com cantos arredondados no Canvas."""
        pts = [
            x1+r, y1,   x2-r, y1,
            x2,   y1,   x2,   y1+r,
            x2,   y2-r, x2,   y2,
            x2-r, y2,   x1+r, y2,
            x1,   y2,   x1,   y2-r,
            x1,   y1+r, x1,   y1,
        ]
        return self.create_polygon(pts, smooth=True, **kwargs)
    
    def _set_hover(self, active):
        if self._state == "disabled": return
        self._state = "hover" if active else "normal"
        self._redraw()
    
    def _set_pressed(self, active):
        if self._state == "disabled": return
        self._state = "pressed" if active else "normal"
        self._redraw()
    
    def _on_release(self, event):
        if self._state == "disabled": return
        self._state = "hover"
        self._redraw()
        if self._command:
            self._command()
    
    def configure(self, state=None, text=None, **kwargs):
        """Compatível com .configure(state='disabled', text='...')"""
        if state == "disabled":
            self._state = "disabled"
            self["cursor"] = ""
        elif state == "normal":
            self._state = "normal"
            self["cursor"] = "hand2"
        if text is not None:
            self._text = text
        self._redraw()
    
    def cget(self, key):
        if key == "state": return self._state
        return super().cget(key)


# ==========================================
# INTERFACE GRÁFICA (GUI CustomTkinter)
# ==========================================

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class AppExtratorPDF(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Apex Extractor Pro - Inteligência de Extração em Massa")
        self.geometry("650x550")
        self.resizable(False, False)
        
        # === Carga da Logo Personalizada (Titlebar + Barra de Tarefas) ===
        try:
            bundle_dir = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
            icon_path = os.path.join(bundle_dir, "logo_extrator.ico")
            self.iconbitmap(icon_path)
            
            # Força o Windows a entender que este App tem uma identidade única na Barra de Tarefas
            myappid = 'antigravity.apexextractor.pro.1.0'
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
            
        except Exception as e:
            pass # Se não achar o arquivo de ícone no ambiente local/empacotado, silencia
            
        self.config_data = load_config()
        self.is_running = False
        self.todos_projetos = []
        self._lock = threading.Lock()
        self._rate_limit_event = threading.Event()
        self._rate_limit_event.set()
        self._log_queue = queue.Queue()  # R2: Fila thread-safe para logs
        self._last_excel_path = None  # U5: Caminho do último Excel gerado
        
        self.mostrar_tela_login()
        self._poll_log_queue()  # Inicia polling de logs na main thread

    # --- TELA DE LICENÇA (AUTENTICAÇÃO) ---
    def mostrar_tela_login(self):
        self.login_frame = ctk.CTkFrame(self, corner_radius=15)
        self.login_frame.pack(pady=80, padx=60, fill="both", expand=True)
        
        title_label = ctk.CTkLabel(self.login_frame, text="�️ Apex Extractor Pro", font=ctk.CTkFont(size=24, weight="bold"))
        title_label.pack(pady=(40, 20))
        
        subtitle = ctk.CTkLabel(self.login_frame, text="Para prosseguir, informe sua Licença Comercial.", font=ctk.CTkFont(size=12), text_color="gray")
        subtitle.pack(pady=(0, 30))
        
        self.lic_entry = ctk.CTkEntry(self.login_frame, placeholder_text="XXXXXXXXXXXXXXXX...", width=300, show="*")
        self.lic_entry.pack(pady=10)
        
        # Carrega licença antiga caso exista
        if self.config_data.get("license_key"):
            self.lic_entry.insert(0, self.config_data["license_key"])

        self.login_btn = ctk.CTkButton(self.login_frame, text="Autenticar", command=self.fazer_login, width=300, height=40)
        self.login_btn.pack(pady=20)
        
        self.login_status = ctk.CTkLabel(self.login_frame, text="", text_color="red")
        self.login_status.pack()

    def fazer_login(self):
        chave = self.lic_entry.get().strip()
        if not chave:
            self.login_status.configure(text="Por favor, digite a Chave de Licença.", text_color="red")
            return
            
        self.login_btn.configure(state="disabled", text="Validando...")
        self.update()
        
        sucesso, mensagem = validate_license(chave)
        
        self.login_btn.configure(state="normal", text="Autenticar")
        
        if sucesso:
            # Salva pra sempre na config e vai pra próxima tela
            self.config_data["license_key"] = chave
            save_config(self.config_data)
            
            messagebox.showinfo("Sucesso", mensagem)
            self.login_frame.destroy()
            self.mostrar_tela_principal()
        else:
            self.login_status.configure(text=mensagem, text_color="red")


    # --- TELA PRINCIPAL DO EXTRATOR ---
    def mostrar_tela_principal(self):
        # Frame de Configuração da API
        api_frame = ctk.CTkFrame(self, fg_color="transparent")
        api_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        provider_frame = ctk.CTkFrame(api_frame, fg_color="transparent")
        provider_frame.pack(fill="x", pady=(0, 5))
        ctk.CTkLabel(provider_frame, text="🤖 Provedor de IA:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=(0, 10))
        
        self.provider_var = ctk.StringVar(value=self.config_data.get("ai_provider", "Gemini"))
        self.provider_menu = ctk.CTkOptionMenu(provider_frame, values=["Gemini", "OpenAI"], variable=self.provider_var, command=self.on_provider_change)
        self.provider_menu.pack(side="left")
        
        self.lbl_api = ctk.CTkLabel(api_frame, text="🔑 Chave de API:", font=ctk.CTkFont(weight="bold"))
        self.lbl_api.pack(anchor="w")
        
        key_container = ctk.CTkFrame(api_frame, fg_color="transparent")
        key_container.pack(fill="x", pady=(5, 0))
        
        self.api_entry = ctk.CTkEntry(key_container, show="*", width=480)
        self.api_entry.pack(side="left")
        
        self.btn_alterar_chave = ctk.CTkButton(key_container, text="Salvar Ativa", command=self.salvar_chave_api, width=100)
        self.btn_alterar_chave.pack(side="right")
        
        # Chamada inicial para preencher a chave correta
        self.on_provider_change(self.provider_var.get())

        # Frame PDF
        pdf_frame = ctk.CTkFrame(self, fg_color="transparent")
        pdf_frame.pack(fill="x", padx=20, pady=(10, 5))
        
        pdf_header = ctk.CTkFrame(pdf_frame, fg_color="transparent")
        pdf_header.pack(fill="x")
        ctk.CTkLabel(pdf_header, text="📄 PDF para Extração:", font=ctk.CTkFont(weight="bold")).pack(side="left")
        
        # U4: Pular primeiras páginas
        ctk.CTkLabel(pdf_header, text="Pular págs:", font=ctk.CTkFont(size=11), text_color="gray").pack(side="right", padx=(10, 5))
        self.skip_pages_var = ctk.StringVar(value="0")
        self.skip_pages_entry = ctk.CTkEntry(pdf_header, textvariable=self.skip_pages_var, width=45, justify="center")
        self.skip_pages_entry.pack(side="right")
        
        pdf_container = ctk.CTkFrame(pdf_frame, fg_color="transparent")
        pdf_container.pack(fill="x", pady=(5, 0))
        
        self.pdf_path_var = ctk.StringVar()
        self.pdf_entry = ctk.CTkEntry(pdf_container, textvariable=self.pdf_path_var, width=480, state="disabled")
        self.pdf_entry.pack(side="left")
        
        ctk.CTkButton(pdf_container, text="Procurar...", command=self.escolher_pdf, width=100).pack(side="right")

        # Frame de Logs
        log_frame = ctk.CTkFrame(self)
        log_frame.pack(fill="both", expand=True, padx=20, pady=(5, 10))
        
        # U1: Barra de progresso com ETA
        progress_header = ctk.CTkFrame(log_frame, fg_color="transparent")
        progress_header.pack(fill="x", padx=10, pady=(10, 2))
        self.eta_label = ctk.CTkLabel(progress_header, text="", font=ctk.CTkFont(size=11), text_color="gray")
        self.eta_label.pack(side="right")
        
        self.progress_bar = ctk.CTkProgressBar(log_frame, mode="determinate")
        self.progress_bar.pack(fill="x", padx=10, pady=(0, 5))
        self.progress_bar.set(0)
        
        self.log_text = ctk.CTkTextbox(log_frame, state="disabled", fg_color="transparent")
        self.log_text.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        # R1: Botão 3D Metálico (com função de cancela/inicia)
        self.start_btn = MetalButton3D(self, text="▶  INICIAR EXTRAÇÃO AUTOMÁTICA",
                                       command=self._toggle_extracao, height=54)
        self.start_btn.pack(fill="x", padx=20, pady=(0, 20))

    def on_provider_change(self, choice):
        self.config_data["ai_provider"] = choice
        save_config(self.config_data)
        
        self.lbl_api.configure(text=f"🔑 {choice} API Key:")
        
        self.api_entry.configure(state="normal")
        self.api_entry.delete(0, "end")
        
        key_name = "gemini_api_key" if choice == "Gemini" else "openai_api_key"
        saved_key = self.config_data.get(key_name, "")
        
        if saved_key:
            self.api_entry.insert(0, saved_key)
            self.api_entry.configure(state="disabled")
            self.btn_alterar_chave.configure(text="Alterar", fg_color="gray", hover_color="darkgray")
        else:
            self.btn_alterar_chave.configure(text="Salvar Ativa", fg_color=["#3a7ebf", "#1f538d"], hover_color=["#325882", "#14375e"])

    def salvar_chave_api(self):
        estado_atual = self.api_entry.cget("state")
        if estado_atual == "disabled":
            self.api_entry.configure(state="normal")
            self.btn_alterar_chave.configure(text="Salvar", fg_color="#3b82f6", hover_color="#2563eb")
        else:
            nova_chave = self.api_entry.get().strip()
            if nova_chave:
                provider = self.provider_var.get()
                key_name = "gemini_api_key" if provider == "Gemini" else "openai_api_key"
                self.config_data[key_name] = nova_chave
                save_config(self.config_data)
                self.api_entry.configure(state="disabled")
                self.btn_alterar_chave.configure(text="Alterar", fg_color="gray", hover_color="darkgray")
                messagebox.showinfo("Salvo", f"Chave da {provider} salva com sucesso!")
            else:
                messagebox.showerror("Erro", "A chave não pode estar vazia.")

    def log(self, message):
        """R2: Thread-safe logging via queue. U3: Timestamps automáticos."""
        ts = datetime.now().strftime("%H:%M:%S")
        self._log_queue.put(f"[{ts}] {message}")
    
    def _poll_log_queue(self):
        """Processa fila de logs na main thread (seguro para Tkinter)."""
        try:
            while True:
                msg = self._log_queue.get_nowait()
                self.log_text.configure(state="normal")
                self.log_text.insert("end", msg + "\n")
                self.log_text.see("end")
                self.log_text.configure(state="disabled")
        except queue.Empty:
            pass
        self.after(100, self._poll_log_queue)  # Polling a cada 100ms
    
    def _toggle_extracao(self):
        """R1: Alterna entre iniciar e cancelar extração."""
        if self.is_running:
            self.is_running = False
            self.log("⛔ Cancelamento solicitado... aguarde os workers finalizarem.")
            self.start_btn.configure(text="⏳ CANCELANDO...")
        else:
            self.iniciar_extracao_thread()
        
    def escolher_pdf(self):
        filepath = filedialog.askopenfilename(
            title="Selecione um PDF",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if filepath:
            self.pdf_path_var.set(filepath)

    def iniciar_extracao_thread(self):
        if self.api_entry.cget("state") == "normal":
             messagebox.showwarning("Aviso", "Por favor, Salve sua Chave de API clicando no botão ao lado antes de continuar.")
             return
             
        provider = self.provider_var.get()
        key_name = "gemini_api_key" if provider == "Gemini" else "openai_api_key"
        api_key_salva = self.config_data.get(key_name, "").strip()
        
        if not api_key_salva:
            messagebox.showwarning("Aviso", f"Por favor, insira a Chave de API da {provider}.")
            return
            
        if not self.pdf_path_var.get().strip():
            messagebox.showwarning("Aviso", "Por favor, selecione um arquivo PDF.")
            return

        # === Trava Militar Anti-Malandragem ===
        self.start_btn.configure(state="disabled", text="🛡️ Verificando Licença Criptografada...")
        self.update()
        
        chave_licenca = self.config_data.get("license_key", "")
        sucesso, msg_licenca = validate_license(chave_licenca)
        if not sucesso:
            messagebox.showerror("Acesso Revogado", f"Acesso Negado: Sua licença foi cancelada, suspensa ou não existe mais no servidor.\nO aplicativo será encerrado de forma forçada imediatamente.\n\nMotivo do Banco de Dados: {msg_licenca}")
            sys.exit(0)
            
        # R1: Botão vira "Cancelar" durante extração
        self.start_btn.configure(state="normal", text="⏹ CANCELAR EXTRAÇÃO")
        self.is_running = True
        
        self.log_text.configure(state="normal")
        self.log_text.delete("0.0", "end")
        self.log_text.configure(state="disabled")
        
        self.progress_bar.set(0)
        self.eta_label.configure(text="")
        self.todos_projetos = []
        self._last_excel_path = None
        
        t = threading.Thread(target=self.processar_pdf, args=(api_key_salva, provider))
        t.daemon = True
        t.start()

    def _worker_fatia(self, fatia_idx, total_fatias, start_com_overlap, end_page, pdf_path, api_key, provider):
        """Worker thread: extrai dados de uma fatia do PDF com retry e backoff.
        P2: Texto lido com list+join. M1: Normaliza emails fragmentados."""
        try:
            doc = fitz.open(pdf_path)
            # P2: Concatenação otimizada via list+join
            parts = []
            for page_num in range(start_com_overlap, end_page + 1):
                parts.append(doc.load_page(page_num).get_text("text"))
            doc.close()
            texto_fatia = "\n\n".join(parts)
            
            # M1: Normaliza emails fragmentados (espaço antes do @)
            texto_fatia = re.sub(r'(\S)\s+@', r'\1@', texto_fatia)
            texto_fatia = re.sub(r'@\s+(\S)', r'@\1', texto_fatia)
            
            if not texto_fatia.strip():
                self.log(f"  [Bloco {fatia_idx}/{total_fatias}] Págs {start_com_overlap+1}-{end_page+1}: vazio, pulando.")
                return (fatia_idx, [])
            
            max_retries = 5
            for tentativa in range(max_retries):
                if not self.is_running:
                    return (fatia_idx, [])
                
                self._rate_limit_event.wait()
                
                try:
                    self.log(f"  [Bloco {fatia_idx}/{total_fatias}] Págs {start_com_overlap+1}-{end_page+1} → {provider} (tentativa {tentativa+1}/{max_retries})...")
                    if provider == "Gemini":
                        # Rate limiting proativo: semafóro + spacing entre releases
                        with _api_semaphore:
                            tabela_markdown = extract_from_gemini(texto_fatia, api_key)
                            time.sleep(_API_SPACING)
                    else:
                        tabela_markdown = extract_from_openai(texto_fatia, api_key)
                    
                    # Detecta se a resposta foi truncada pelo limite de tokens
                    if "<!-- TRUNCADO:" in tabela_markdown:
                        self.log(f"  [Bloco {fatia_idx}/{total_fatias}] ⚠️ Resposta TRUNCADA pelo limite de tokens — dados podem estar incompletos")

                    novos_projetos = parse_markdown_table_to_dicts(tabela_markdown)
                    if novos_projetos:
                        expandidos = expandir_por_autor(novos_projetos)
                        self.log(f"  [Bloco {fatia_idx}/{total_fatias}] ✅ {len(novos_projetos)} trabalhos → {len(expandidos)} linhas")
                        return (fatia_idx, expandidos)
                    else:
                        self.log(f"  [Bloco {fatia_idx}/{total_fatias}] Sem resultados neste trecho.")
                        return (fatia_idx, [])
                    
                except Exception as e:
                    erro_msg = str(e)
                    if tentativa < max_retries - 1:
                        is_rate_limit = ("429" in erro_msg or "quota" in erro_msg.lower() or 
                                        "too many requests" in erro_msg.lower() or 
                                        "rate_limit_exceeded" in erro_msg.lower() or
                                        "insufficient_quota" in erro_msg.lower())
                        if is_rate_limit:
                            wait_time = 60
                            match = re.search(r"retry in ([\d\.]+)s", erro_msg)
                            if match:
                                wait_time = int(float(match.group(1))) + 5
                            else:
                                match_json = re.search(r'"retryDelay":\s*"(\d+)s"', erro_msg)
                                if match_json:
                                    wait_time = int(match_json.group(1)) + 5
                                else:
                                    match_openai = re.search(r"Please try again in (\d+(\.\d+)?)s", erro_msg)
                                    if match_openai:
                                        wait_time = int(float(match_openai.group(1))) + 5
                            
                            self.log(f"  [Bloco {fatia_idx}] ⚠️ Rate-limit 429! Travando TODOS os workers por {wait_time}s...")
                            self._rate_limit_event.clear()
                            time.sleep(wait_time)
                            self._rate_limit_event.set()
                        else:
                            self.log(f"  [Bloco {fatia_idx}] Erro API (tentativa {tentativa+1}): {erro_msg[:120]}")
                            time.sleep(5)
                    else:
                        self.log(f"  [Bloco {fatia_idx}/{total_fatias}] ❌ FALHA APÓS {max_retries} TENTATIVAS: {erro_msg[:120]}")
            
            return (fatia_idx, None)
        except Exception as e:
            self.log(f"  [Bloco {fatia_idx}] ERRO WORKER: {str(e)[:120]}")
            return (fatia_idx, None)

    def processar_pdf(self, api_key, provider):
        try:
            pdf_path = self.pdf_path_var.get()
            chunk_size  = 12   # 12 págs/bloco → lotes menores para garantir que a IA não sofra de 'preguiça' e extraia TUDO
            overlap     = 3
            
            # U4: Pular primeiras páginas
            try:
                skip_pages = int(self.skip_pages_var.get())
            except ValueError:
                skip_pages = 0
            
            self.log(f"Abrindo arquivo PDF...")
            doc = fitz.open(pdf_path)
            total_pages = len(doc)
            doc.close()
            
            if skip_pages > 0:
                self.log(f"⏩ Pulando as primeiras {skip_pages} páginas (capa/sumário).")
            
            # Calcula fatias a partir de skip_pages
            fatias = []
            for start_page in range(skip_pages, total_pages, chunk_size):
                start_com_overlap = max(skip_pages, start_page - overlap)
                end_page = min(start_page + chunk_size, total_pages) - 1
                fatias.append((start_com_overlap, end_page))
            
            total_fatias = len(fatias)
            pages_efetivas = total_pages - skip_pages
            
            self.log(f"⚡ MODO TURBO PARALELO ATIVADO")
            # Loga o modelo Gemini selecionado para transparência
            if provider == "Gemini":
                try:
                    modelo = _get_gemini_model(self.config_data.get("gemini_api_key", ""))
                    self.log(f"🤖 Modelo selecionado: {modelo}")
                except Exception:
                    pass
            self.log(f"Total: {total_pages} págs ({pages_efetivas} efetivas) | {total_fatias} blocos | 3 workers | Motor: {provider}")
            
            base_name = os.path.splitext(os.path.basename(pdf_path))[0]
            base_dir = os.path.dirname(pdf_path)
            excel_filename = os.path.join(base_dir, f"{base_name}_Extraido.xlsx")
            self._last_excel_path = excel_filename
            
            # ========== FASE 1: Disparo paralelo ==========
            concluidas = 0
            fatias_falhadas = []
            time_start = time.time()  # U1: Cronômetro para ETA
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                futures = {}
                for i, (s, e) in enumerate(fatias):
                    f = executor.submit(self._worker_fatia, i+1, total_fatias, s, e, pdf_path, api_key, provider)
                    futures[f] = (i + 1, s, e)
                
                for future in concurrent.futures.as_completed(futures):
                    if not self.is_running:
                        self.log("⛔ Processo cancelado pelo usuário!")
                        executor.shutdown(wait=False, cancel_futures=True)
                        break
                    
                    fatia_num, s, e = futures[future]
                    fatia_idx, resultados_fatia = future.result()
                    concluidas += 1
                    
                    # U1: Calcula ETA
                    elapsed = time.time() - time_start
                    if concluidas > 0:
                        avg_per_block = elapsed / concluidas
                        remaining = (total_fatias - concluidas) * avg_per_block
                        mins, secs = divmod(int(remaining), 60)
                        self.eta_label.configure(text=f"⏱ ~{mins}m{secs:02d}s restantes")
                    
                    if resultados_fatia is None:
                        fatias_falhadas.append((fatia_idx, s, e))
                    elif resultados_fatia:
                        with self._lock:
                            self.todos_projetos.extend(resultados_fatia)
                            self.todos_projetos = consolidar_projetos(self.todos_projetos)
                            
                            # P3: Debounce — salva Excel a cada 3 blocos (não a cada 1)
                            if concluidas % 3 == 0 or concluidas == total_fatias:
                                try:
                                    salvar_excel_formatado(self.todos_projetos, excel_filename)
                                except PermissionError as pe:
                                    self.log(f"  -> AVISO: {str(pe)}")
                                except Exception as ee:
                                    self.log(f"  -> ERRO ao salvar Excel: {str(ee)}")
                    
                    self.progress_bar.set(concluidas / float(total_fatias))
            
            # ========== FASE 2: Retentativa sequencial ==========
            if fatias_falhadas and self.is_running:
                self.log(f"\n🔄 FASE 2: Retentando {len(fatias_falhadas)} blocos que falharam (modo sequencial)...")
                for fatia_idx, s, e in fatias_falhadas:
                    if not self.is_running:
                        break
                    self.log(f"  Re-processando bloco {fatia_idx} (págs {s+1}-{e+1})...")
                    _, resultados = self._worker_fatia(fatia_idx, total_fatias, s, e, pdf_path, api_key, provider)
                    if resultados:
                        with self._lock:
                            self.todos_projetos.extend(resultados)
                            self.todos_projetos = consolidar_projetos(self.todos_projetos)
                            self.log(f"  [Bloco {fatia_idx}] 🔁 Recuperado! +{len(resultados)} linhas")
                            try:
                                salvar_excel_formatado(self.todos_projetos, excel_filename)
                            except Exception:
                                pass
                    elif resultados is None:
                        self.log(f"  [Bloco {fatia_idx}] ❌ Falha definitiva mesmo na retentativa.")
            
            # ========== Salvamento final e relatório ==========
            if self.is_running and self.todos_projetos:
                # Salvamento final garantido
                try:
                    salvar_excel_formatado(self.todos_projetos, excel_filename)
                except Exception:
                    pass
                
                elapsed_total = time.time() - time_start
                mins, secs = divmod(int(elapsed_total), 60)
                emails_extraidos = sum(1 for p in self.todos_projetos if p.get('E-mails dos Autores', '').strip())
                
                self.eta_label.configure(text=f"✅ Concluído em {mins}m{secs:02d}s")
                self.log(f"\n🎉 PROCESSO CONCLUÍDO 100%!")
                self.log(f"📊 Relatório: {len(self.todos_projetos)} linhas | {emails_extraidos} com e-mail | Tempo: {mins}m{secs:02d}s")
                if fatias_falhadas:
                    self.log(f"⚠️ {len(fatias_falhadas)} blocos tiveram problemas (retentados na Fase 2).")
                self.log(f"Planilha salva em: {excel_filename}")
                
                # U5: Pergunta se quer abrir a planilha
                abrir = messagebox.askyesno("Sucesso", 
                    f"Extração concluída!\n\n"
                    f"• {len(self.todos_projetos)} linhas totais\n"
                    f"• {emails_extraidos} com e-mail\n"
                    f"• Tempo: {mins}m{secs:02d}s\n\n"
                    f"Deseja abrir a planilha agora?")
                if abrir:
                    os.startfile(excel_filename)
            elif not self.todos_projetos and self.is_running:
                self.log("\n⚠️ Nenhum dado extraído. Verifique se o PDF contém trabalhos acadêmicos.")
                messagebox.showwarning("Aviso", "Nenhum dado foi extraído do PDF.")
            
        except Exception as e:
            self.log(f"\nERRO FATAL: {str(e)}")
            messagebox.showerror("Erro Crítico", f"Ocorreu um erro extremo:\n\n{str(e)}")
        finally:
            self.start_btn.configure(state="normal", text="▶ INICIAR EXTRAÇÃO AUTOMÁTICA")
            self.is_running = False

# Inicialização
if __name__ == "__main__":
    app = AppExtratorPDF()
    app.mainloop()
