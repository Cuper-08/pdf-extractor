"""
Script de teste end-to-end do PDF Extractor.
Envia o PDF para a API Python, verifica os chunks no Supabase
e valida os dados extraídos na tabela extracted_projects.
"""
import requests
import time
import json
import sys
import os

# ── Configurações ────────────────────────────────────────────────────────────
# Servidor local (rode `uvicorn main:app --port 8001` antes de executar este script)
API_URL = "http://localhost:8001"

SUPABASE_URL = "https://zcoixlvbkssvuxamzwvs.supabase.co"
SUPABASE_ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inpjb2l4bHZia3NzdnV4YW16d3ZzIiwicm9sZSI6"
    "ImFub24iLCJpYXQiOjE3NjkzOTYyNzIsImV4cCI6MjA4NDk3MjI3Mn0"
    ".BVPwqGMc7RhqRSOjnqIsZjwhe4Ls9pf9aajX-by-98M"
)
HEADERS = {
    "apikey": SUPABASE_ANON_KEY,
    "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
}

PDF_PATH = os.path.join(os.path.dirname(__file__), "tests", "Medio (1).pdf")


def separator(label):
    print(f"\n{'='*60}")
    print(f"  {label}")
    print('='*60)


def step1_health():
    separator("STEP 1: Health Check da API")
    try:
        r = requests.get(f"{API_URL}/health", timeout=10)
        print(f"  Status: {r.status_code}")
        print(f"  Body:   {r.json()}")
        if r.status_code == 200:
            print("  ✅ API respondendo corretamente")
            return True
        else:
            print("  ❌ API não está saudável")
            return False
    except Exception as e:
        print(f"  ❌ Erro ao chamar /health: {e}")
        return False


def step2_upload_pdf():
    separator("STEP 2: Upload do PDF para /extract")
    if not os.path.exists(PDF_PATH):
        print(f"  ❌ PDF não encontrado em: {PDF_PATH}")
        return None

    size_mb = os.path.getsize(PDF_PATH) / (1024 * 1024)
    print(f"  Arquivo: {os.path.basename(PDF_PATH)}")
    print(f"  Tamanho: {size_mb:.2f} MB")
    print(f"  Enviando para {API_URL}/extract ...")

    try:
        with open(PDF_PATH, "rb") as f:
            r = requests.post(
                f"{API_URL}/extract",
                files={"file": (os.path.basename(PDF_PATH), f, "application/pdf")},
                timeout=300,
            )

        print(f"  HTTP Status: {r.status_code}")

        if r.status_code in (200, 202):
            data = r.json()
            print(f"  ✅ Aceito!")
            print(f"     document_id  : {data.get('document_id')}")
            print(f"     filename     : {data.get('filename')}")
            print(f"     total_pages  : {data.get('total_pages')}")
            print(f"     total_chunks : {data.get('total_chunks')}")
            return data.get("document_id")
        else:
            print(f"  ❌ Erro HTTP {r.status_code}")
            try:
                print(f"  Body: {json.dumps(r.json(), indent=2, ensure_ascii=False)}")
            except Exception:
                print(f"  Body raw: {r.text[:500]}")
            return None

    except Exception as e:
        print(f"  ❌ Exceção ao enviar: {e}")
        return None


def step3_check_chunks(document_id):
    separator("STEP 3: Verificando chunks no Supabase")
    url = (
        f"{SUPABASE_URL}/rest/v1/pdf_chunks"
        f"?document_id=eq.{document_id}&select=id,chunk_index,processing_status,estimated_pages"
        f"&order=chunk_index.asc"
    )
    r = requests.get(url, headers=HEADERS, timeout=15)
    if r.status_code != 200:
        print(f"  ❌ Erro ao consultar Supabase: {r.status_code} — {r.text[:300]}")
        return 0

    chunks = r.json()
    print(f"  Total de chunks criados: {len(chunks)}")
    for c in chunks[:5]:
        print(f"    Chunk #{c['chunk_index']:3d} | Status: {c['processing_status']:10s} | Páginas: {c['estimated_pages']}")
    if len(chunks) > 5:
        print(f"    ... e mais {len(chunks)-5} chunks")
    return len(chunks)


def step4_wait_worker(document_id, total_chunks, max_wait=300):
    separator("STEP 4: Aguardando o WORKER processar os chunks")
    print(f"  (Verificando a cada 15s por até {max_wait}s)")
    elapsed = 0
    while elapsed < max_wait:
        url = (
            f"{SUPABASE_URL}/rest/v1/pdf_chunks"
            f"?document_id=eq.{document_id}&processing_status=eq.processed&select=id"
        )
        r = requests.get(url, headers=HEADERS, timeout=15)
        processed = len(r.json()) if r.status_code == 200 else 0
        pct = (processed / total_chunks * 100) if total_chunks > 0 else 0
        print(f"  [{elapsed:3d}s] Processados: {processed}/{total_chunks} ({pct:.0f}%)")
        if processed >= total_chunks:
            print("  ✅ Todos os chunks processados!")
            return True
        time.sleep(15)
        elapsed += 15
    print("  ⚠️  Tempo esgotado — verificando resultados parciais")
    return False


def step5_check_results(document_id):
    separator("STEP 5: Verificando resultados na tabela extracted_projects")
    url = (
        f"{SUPABASE_URL}/rest/v1/extracted_projects"
        f"?document_id=eq.{document_id}&select=project_title,authors,emails"
        f"&order=created_at.asc"
    )
    r = requests.get(url, headers=HEADERS, timeout=15)
    if r.status_code != 200:
        print(f"  ❌ Erro ao consultar Supabase: {r.status_code}")
        return []

    results = r.json()
    # Filtra apenas os que têm título (descarta chunks sem dados)
    valid = [x for x in results if x.get("project_title")]
    print(f"  Total de registros: {len(results)} ({len(valid)} com título)")
    print()
    for i, row in enumerate(valid[:10], 1):
        print(f"  📄 Projeto {i}:")
        print(f"     Título  : {row.get('project_title', '')[:80]}")
        print(f"     Autores : {row.get('authors', '')[:80]}")
        print(f"     E-mails : {row.get('emails', '')[:80]}")
        print()
    if len(valid) > 10:
        print(f"  ... e mais {len(valid)-10} projetos")
    return valid


def step6_export_excel(results, document_id):
    separator("STEP 6: Exportando para Excel")
    try:
        import openpyxl
    except ImportError:
        print("  ⚠️  openpyxl não instalado — instalando agora...")
        os.system("pip install openpyxl -q")
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projetos Extraídos"

    # Cabeçalho
    ws["A1"] = "Título do Trabalho"
    ws["B1"] = "Autores"
    ws["C1"] = "E-mails"
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, proj in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=proj.get("project_title", ""))
        ws.cell(row=row_idx, column=2, value=proj.get("authors", ""))
        ws.cell(row=row_idx, column=3, value=proj.get("emails", ""))

    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40

    output_path = os.path.join(
        os.path.dirname(__file__), "tests", f"resultado_{document_id[:8]}.xlsx"
    )
    wb.save(output_path)
    print(f"  ✅ Excel salvo em: {output_path}")
    print(f"     {len(results)} projetos exportados")
    return output_path


if __name__ == "__main__":
    print("\n🚀 TESTE END-TO-END — PDF Extractor Assíncrono")

    # 1. Health
    if not step1_health():
        print("\n❌ API offline. Verifique o Coolify.")
        sys.exit(1)

    # 2. Upload
    doc_id = step2_upload_pdf()
    if not doc_id:
        print("\n❌ Upload falhou. Verifique os logs do Coolify.")
        sys.exit(1)

    # 3. Chunks
    total = step3_check_chunks(doc_id)
    if total == 0:
        print("\n❌ Nenhum chunk foi criado no Supabase.")
        sys.exit(1)

    # 4. Worker (aguardar)
    step4_wait_worker(doc_id, total)

    # 5. Resultados
    results = step5_check_results(doc_id)

    # 6. Excel
    valid = [r for r in results if r.get("project_title")]
    if valid:
        step6_export_excel(valid, doc_id)
    else:
        print("\n⚠️  Nenhum projeto com dados completos foi extraído ainda.")
        print("    O WORKER pode ainda estar processando. Tente novamente em alguns minutos.")

    print("\n✅ Teste concluído!")
