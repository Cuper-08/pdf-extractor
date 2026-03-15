"""
Motor de extração do portal — 2 camadas de IA.

Layer 1 — generate_schema():
  Recebe descrição em linguagem natural → Gemini gera system_prompt customizado
  + lista de colunas (JSON). Roda UMA VEZ por job.

Layer 2 — process_pdf_extraction():
  PDF → PyMuPDF → chunks (smart_split) → Gemini processa em paralelo
  → retorna JSON → dedup → retorna list[dict]

Funções utilitárias:
  smart_split, consolidar_generico, create_excel
"""
import os
import re
import json
import asyncio
import unicodedata
from io import BytesIO
from typing import Optional, Callable, Awaitable

import fitz  # PyMuPDF
import httpx
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
CHUNK_SIZE = 40_000        # chars por chunk (~10 páginas densas)
CHUNK_OVERLAP = 2_000      # overlap entre chunks para não perder registros no limite
MAX_CONCURRENT = 3         # máximo de chunks em paralelo (evita rate limit)

# ─────────────────────────────────────────────
# GEMINI — model discovery (async, cached)
# ─────────────────────────────────────────────

_model_cache: dict[str, str] = {}
_model_lock = asyncio.Lock()

PREFERRED_MODELS = [
    "models/gemini-1.5-flash",        # Rápido, estável — ideal para extração
    "models/gemini-1.5-pro",          # Mais poderoso
]


async def _get_gemini_model(api_key: str) -> str:
    if api_key in _model_cache:
        return _model_cache[api_key]

    async with _model_lock:
        if api_key in _model_cache:
            return _model_cache[api_key]

        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}",
                timeout=15,
            )
        if resp.status_code != 200:
            raise Exception(f"Chave Gemini inválida (HTTP {resp.status_code})")

        # Modelos descontinuados que a API ainda lista mas rejeitam requests
        DEPRECATED = {"models/gemini-2.0-flash", "models/gemini-2.0-flash-001", "models/gemini-pro"}
        available = [m["name"] for m in resp.json().get("models", [])
                     if m["name"] not in DEPRECATED]
        model = None
        for pref in PREFERRED_MODELS:
            if pref in available:
                model = pref
                break
        if not model:
            flash = [m for m in available if "flash" in m.lower() and "001" not in m]
            model = flash[0] if flash else available[0]

        _model_cache[api_key] = model
        return model


async def _call_gemini(system_prompt: str, user_text: str, api_key: str, response_json: bool = False) -> str:
    model = await _get_gemini_model(api_key)
    url = f"https://generativelanguage.googleapis.com/v1beta/{model}:generateContent?key={api_key}"
    payload = {
        "systemInstruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"parts": [{"text": user_text}]}],
        "generationConfig": {"temperature": 0.0},
    }
    if response_json:
        payload["generationConfig"]["responseMimeType"] = "application/json"

    async with httpx.AsyncClient() as client:
        resp = await client.post(url, json=payload, timeout=60)

    # Se modelo retornou 404 (descontinuado), limpa cache e tenta redescobrir
    if resp.status_code == 404:
        _model_cache.pop(api_key, None)
        model = await _get_gemini_model(api_key)
        url = f"https://generativelanguage.googleapis.com/v1beta/{model}:generateContent?key={api_key}"
        async with httpx.AsyncClient() as client:
            resp = await client.post(url, json=payload, timeout=60)

    # Retry logic com backoff em 429
    if resp.status_code == 429:
        for wait in (15, 30):
            await asyncio.sleep(wait)
            async with httpx.AsyncClient() as client:
                resp = await client.post(url, json=payload, timeout=60)
            if resp.status_code != 429:
                break

    if resp.status_code != 200:
        raise Exception(f"Gemini error {resp.status_code}: {resp.text[:500]}")

    data = resp.json()
    try:
        return data["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError):
        raise Exception(f"Resposta inesperada do Gemini: {json.dumps(data)[:300]}")


# ─────────────────────────────────────────────
# LAYER 1 — generate_schema
# ─────────────────────────────────────────────

META_SYSTEM = """Você é um especialista em extração de dados estruturados de documentos PDF.
Sua resposta deve ser OBRIGATORIAMENTE um JSON válido estruturado de acordo com as instruções.
"""

META_USER_TEMPLATE = """O usuário quer extrair as seguintes informações de seus documentos PDF:

{user_request}

Crie um system prompt preciso para que uma IA extraia esses dados documentais com máxima precisão.

Retorne EXATAMENTE este objeto JSON de raiz (não use markdown code blocks se o mime-type já for JSON):
{{
  "system_prompt": "...",
  "columns": ["coluna1", "coluna2", ...]
}}

O system_prompt OBRIGATORIAMENTE deve:
1. Definir exatamente quais campos extrair.
2. Exigir que a saída seja EXCLUSIVAMENTE um array de objetos JSON `[{{ "coluna1": "valor" }}]`, onde as chaves sejam EXATAMENTE as colunas definidas.
3. Instruir a deixar valores VAZIOS (string vazia ou null) quando o dado não existir (NUNCA inventar dados).
4. Definir UMA LINHA/OBJETO por registro/item distinto encontrado no documento.
5. Instruir a extrair TODOS os registros encontrados no texto lido, sem pular nenhum.
6. Instruir a repetir campos-chave de contexto em cada objeto filho (ex: o "nome do processo" ou "título do documento" deve ser repetido em cada registro associado).
7. Instruir a ignorar inteiramente cabeçalhos de página, currículos irrelevantes e numeração de página.
8. Ser escrito em português claro e direto.

O campo "columns" abaixo deve listar as colunas exatas que serão usadas como chaves nos objetos JSON extraídos."""


async def generate_schema(user_description: str) -> dict:
    """
    Layer 1: gera system_prompt + colunas a partir da descrição do usuário.
    Retorna: {system_prompt: str, columns: list[str], description: str}
    """
    api_key = GEMINI_API_KEY
    if not api_key:
        raise Exception("GEMINI_API_KEY não configurada no servidor.")

    user_text = META_USER_TEMPLATE.format(user_request=user_description.strip())
    raw = await _call_gemini(META_SYSTEM, user_text, api_key, response_json=True)

    # Parse JSON
    try:
        parsed = json.loads(raw.strip())
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if not match:
            raise Exception(f"Resposta da IA não é JSON válido: {raw[:300]}")
        parsed = json.loads(match.group())

    system_prompt = parsed.get("system_prompt", "")
    columns = parsed.get("columns", [])

    if not system_prompt or not columns:
        raise Exception("Resposta incompleta do gerador de esquema. Tente reformular a descrição.")

    return {
        "system_prompt": system_prompt,
        "columns": columns,
        "description": user_description,
    }


# ─────────────────────────────────────────────
# LAYER 2 — process_pdf_extraction
# ─────────────────────────────────────────────

def smart_split(full_text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """
    Divide texto em chunks nas fronteiras naturais (parágrafo → frase → espaço).
    Inclui overlap entre chunks para garantir que registros no limite não sejam perdidos.
    O dedup em consolidar_generico remove as duplicatas geradas pelo overlap.
    """
    chunks = []
    total = len(full_text)
    pos = 0

    while pos < total:
        end = min(pos + chunk_size, total)

        if end < total:
            bp = full_text.rfind("\n\n", pos, end)
            if bp == -1 or bp < pos + chunk_size // 2:
                bp = full_text.rfind(".\n", pos, end)
            if bp == -1 or bp < pos + chunk_size // 2:
                bp = full_text.rfind(". ", pos, end)
            if bp != -1 and bp > pos + chunk_size // 2:
                end = bp + 1

        chunks.append(full_text[pos:end])
        # Avança mantendo overlap com o próximo chunk; garante progresso mínimo de 1 char
        pos = max(pos + 1, end - overlap)

    return chunks


async def _extract_chunk(chunk: str, schema_prompt: str, sem: asyncio.Semaphore) -> str:
    async with sem:
        return await _call_gemini(
            schema_prompt,
            f"Extraia os dados deste texto:\n\n{chunk}",
            GEMINI_API_KEY,
            response_json=True,
        )


async def process_pdf_extraction(
    pdf_bytes: bytes,
    schema_prompt: str,
    on_progress: Optional[Callable[[int], Awaitable[None]]] = None,
) -> list[dict]:
    """
    Layer 2: extrai texto do PDF → divide em chunks → processa com Gemini
    em paralelo → faz parse → dedup → retorna list[dict].

    on_progress: corrotina opcional chamada após cada chunk (recebe % 0-100).
    """
    # 1. Extrai texto do PDF
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    full_text = "\n".join(page.get_text() for page in doc)
    doc.close()

    if not full_text.strip():
        raise Exception("PDF sem texto extraível. O arquivo pode ser baseado em imagens (scan).")

    # 2. Divide em chunks
    chunks = smart_split(full_text)
    total = len(chunks)
    completed_count = 0

    # 3. Processa em paralelo com semáforo de rate limit + rastreio de progresso
    sem = asyncio.Semaphore(MAX_CONCURRENT)

    async def _extract_with_progress(chunk: str) -> str:
        nonlocal completed_count
        result = await _extract_chunk(chunk, schema_prompt, sem)
        completed_count += 1
        if on_progress:
            pct = min(99, int(completed_count / total * 100))  # 99% até salvar no DB
            try:
                await on_progress(pct)
            except Exception:
                pass  # progresso é best-effort
        return result

    tasks = [_extract_with_progress(chunk) for chunk in chunks]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    # 4. Coleta registros (ignora chunks com erro)
    all_records: list[dict] = []
    for r in results:
        if isinstance(r, Exception):
            continue  # chunk falhou — best effort
            
        try:
            # remove markers markdown code se retornados erradamente
            clean_r = r.strip()
            if clean_r.startswith("```json"):
                clean_r = clean_r[7:]
            if clean_r.endswith("```"):
                clean_r = clean_r[:-3]
                
            records = json.loads(clean_r.strip())
            
            if isinstance(records, list):
                all_records.extend(records)
            elif isinstance(records, dict):
                # Caso a IA retorne algo tipo {"data": [...]} ou só um registro
                if len(records.values()) == 1 and isinstance(list(records.values())[0], list):
                    all_records.extend(list(records.values())[0])
                else:
                    all_records.append(records)
        except json.JSONDecodeError:
            pass  # Ignora chunk com falha no JSON silenciosamente

    # 5. Remove linhas completamente vazias
    all_records = [r for r in all_records if any(str(v).strip() for v in r.values())]

    # 6. Dedup e retorna
    return consolidar_generico(all_records)


def consolidar_generico(records: list[dict]) -> list[dict]:
    """
    Dedup por igualdade exata de todos os campos (NFC normalizado).
    Apenas linhas 100% idênticas são removidas — registros distintos com
    os mesmos valores em algumas colunas (ex: mesmo autor em obras diferentes)
    são preservados normalmente.
    """
    if not records:
        return []

    keys = list(records[0].keys())
    seen: set[str] = set()
    result: list[dict] = []

    for rec in records:
        chave = "|||".join(
            unicodedata.normalize("NFC", str(rec.get(k) or "").strip().lower())
            for k in keys
        )
        if chave not in seen:
            seen.add(chave)
            result.append(rec)

    return result


# ─────────────────────────────────────────────
# EXCEL — geração via BytesIO
# ─────────────────────────────────────────────

def create_excel(records: list[dict], columns: Optional[list[str]] = None) -> bytes:
    """
    Gera Excel formatado a partir de list[dict].
    Retorna bytes (para armazenar no Supabase ou streaming download).
    """
    if records:
        df = pd.DataFrame(records)
        if columns:
            # Reordena pelas colunas esperadas; adiciona as extras no final
            ordered = [c for c in columns if c in df.columns]
            extras = [c for c in df.columns if c not in columns]
            df = df[ordered + extras]
        df = df.fillna("")
    else:
        df = pd.DataFrame(columns=columns or [])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Dados Extraídos")
        ws = writer.sheets["Dados Extraídos"]

        # Auto-largura das colunas
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Estilo do cabeçalho (azul escuro)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    output.seek(0)
    return output.read()
